#!/usr/bin/env python3
"""
Azure Content Safety Batch Processor - Production Ready Enhanced Version
Comprehensive solution for handling multiple dataset schemas with advanced analytics
"""

import asyncio
import aiohttp
import pandas as pd
import json
import time
import logging
import sys
import os
import argparse
import traceback
import re
from pathlib import Path
from datetime import datetime
from typing import List, Dict, Any, Optional, Tuple, Union
from collections import defaultdict, Counter

# Python compatibility
def run_async_main(coro):
    try:
        return asyncio.run(coro)
    except AttributeError:
        loop = asyncio.get_event_loop()
        try:
            return loop.run_until_complete(coro)
        finally:
            loop.close()

# Dependencies
try:
    from sklearn.metrics import (
        confusion_matrix, f1_score, precision_score, recall_score, 
        accuracy_score, classification_report, roc_auc_score
    )
    SKLEARN_AVAILABLE = True
except ImportError:
    SKLEARN_AVAILABLE = False
    print("WARNING: scikit-learn not available. Install with: pip install scikit-learn")

try:
    import numpy as np
    NUMPY_AVAILABLE = True
except ImportError:
    NUMPY_AVAILABLE = False
    print("WARNING: numpy not available. Install with: pip install numpy")

try:
    import matplotlib.pyplot as plt
    import seaborn as sns
    PLOTTING_AVAILABLE = True
except ImportError:
    PLOTTING_AVAILABLE = False

try:
    from openpyxl import Workbook
    from openpyxl.styles import PatternFill, Font
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False
    print("WARNING: openpyxl not available. Excel output disabled. Install with: pip install openpyxl")

# Configure logging
azure_results_dir = Path("azure_results")
azure_results_dir.mkdir(exist_ok=True)

logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    handlers=[
        logging.FileHandler(azure_results_dir / 'azure_content_safety_batch.log'),
        logging.StreamHandler()
    ]
)
logger = logging.getLogger(__name__)

class DatasetSchemaDetector:
    """Advanced dataset schema detection and validation"""
    
    def __init__(self):
        self.known_schemas = {
            'jbb_benign': {
                'name': 'JBB Benign Behaviors',
                'prompt_columns': ['behavior', 'goal'],
                'ground_truth_columns': ['category', 'type'],
                'indicators': ['behavior', 'benign', 'jbb'],
                'description': 'JBB dataset with benign behaviors'
            },
            'jailbreak_balanced': {
                'name': 'Jailbreak Balanced Dataset',
                'prompt_columns': ['jailbreak_query', 'query', 'prompt'],
                'ground_truth_columns': ['selected', 'type', 'attack_type'],
                'indicators': ['jailbreak', 'balanced', 'classification'],
                'description': 'Balanced jailbreak classification dataset'
            },
            'mini_jailbreak': {
                'name': 'Mini JailBreak Dataset',
                'prompt_columns': ['jailbreak_query', 'query'],
                'ground_truth_columns': ['selected_mini', 'transfer_from_llm'],
                'indicators': ['mini', 'jailbreak', '28k'],
                'description': 'Mini jailbreak dataset with transfer indicators'
            },
            'redteam_general': {
                'name': 'Red Team Dataset',
                'prompt_columns': ['prompt', 'query', 'input', 'user_input'],
                'ground_truth_columns': ['harmful', 'attack_type', 'safety_label'],
                'indicators': ['redteam', 'red_team', 'harmful'],
                'description': 'General red team evaluation dataset'
            },
            'safety_eval': {
                'name': 'Safety Evaluation Dataset',
                'prompt_columns': ['prompt', 'text', 'input'],
                'ground_truth_columns': ['safety_label', 'risk_category', 'violation_type'],
                'indicators': ['safety', 'eval', 'evaluation'],
                'description': 'Safety evaluation with risk categories'
            }
        }
    
    def detect_schema(self, df: pd.DataFrame, filename: str = "") -> Dict[str, Any]:
        """Detect dataset schema with confidence scoring"""
        filename_lower = filename.lower()
        columns_lower = [col.lower() for col in df.columns]
        
        schema_scores = {}
        
        for schema_id, schema_info in self.known_schemas.items():
            score = 0
            matches = []
            
            # Filename indicators
            for indicator in schema_info['indicators']:
                if indicator in filename_lower:
                    score += 20
                    matches.append(f"filename:{indicator}")
            
            # Prompt column matches
            prompt_matches = 0
            for prompt_col in schema_info['prompt_columns']:
                if prompt_col in columns_lower:
                    prompt_matches += 1
                    score += 15
                    matches.append(f"prompt_col:{prompt_col}")
            
            # Ground truth column matches
            gt_matches = 0
            for gt_col in schema_info['ground_truth_columns']:
                if gt_col in columns_lower:
                    gt_matches += 1
                    score += 15
                    matches.append(f"gt_col:{gt_col}")
            
            # Bonus for having both prompt and GT columns
            if prompt_matches > 0 and gt_matches > 0:
                score += 10
            
            schema_scores[schema_id] = {
                'score': score,
                'matches': matches,
                'schema_info': schema_info
            }
        
        # Find best match
        best_schema = max(schema_scores.items(), key=lambda x: x[1]['score'])
        
        result = {
            'detected_schema': best_schema[0] if best_schema[1]['score'] > 30 else 'unknown',
            'confidence': min(best_schema[1]['score'] / 100, 1.0),
            'schema_info': best_schema[1]['schema_info'] if best_schema[1]['score'] > 30 else None,
            'all_scores': schema_scores,
            'filename': filename,
            'columns': list(df.columns),
            'sample_size': len(df)
        }
        
        return result

class AdvancedGroundTruthMapper:
    """Enhanced ground truth mapper with ML-ready features"""
    
    def __init__(self):
        # Expanded keyword mappings with weights
        self.malicious_keywords = {
            # High confidence malicious
            'jailbreak': 0.95, 'malicious': 0.95, 'harmful': 0.9, 'attack': 0.9,
            'exploit': 0.9, 'hack': 0.85, 'unsafe': 0.8, 'dangerous': 0.8,
            'toxic': 0.8, 'abusive': 0.8, 'inappropriate': 0.75, 'illegal': 0.9,
            
            # Medium confidence malicious
            'violation': 0.7, 'adversarial': 0.75, 'red_team': 0.8, 'redteam': 0.8,
            'prompt_injection': 0.85, 'bad': 0.6, 'negative': 0.5, 'prohibited': 0.8,
            
            # Context-specific malicious
            'hate': 0.9, 'violence': 0.9, 'sexual': 0.8, 'harassment': 0.85,
            'discrimination': 0.8, 'self_harm': 0.9, 'selfharm': 0.9,
            'terrorism': 0.95, 'extremism': 0.9
        }
        
        self.benign_keywords = {
            # High confidence benign
            'benign': 0.95, 'safe': 0.9, 'normal': 0.85, 'clean': 0.8,
            'good': 0.7, 'positive': 0.7, 'legitimate': 0.85, 'appropriate': 0.8,
            
            # Medium confidence benign
            'acceptable': 0.75, 'allowed': 0.8, 'permitted': 0.8, 'compliant': 0.85,
            'ok': 0.6, 'fine': 0.6, 'valid': 0.7, 'proper': 0.7,
            
            # Context-specific benign
            'creative_writing': 0.8, 'education': 0.8, 'information_seeking': 0.75,
            'help': 0.7, 'assistance': 0.7, 'support': 0.7, 'guidance': 0.7
        }
        
        # Pattern-based classifications
        self.malicious_patterns = [
            (r'.*jailbreak.*', 0.9),
            (r'.*attack.*', 0.85),
            (r'.*exploit.*', 0.85),
            (r'.*harmful.*', 0.8),
            (r'.*unsafe.*', 0.75),
            (r'.*violation.*', 0.7),
            (r'.*prohibited.*', 0.8),
            (r'.*illegal.*', 0.9),
            (r'.*dangerous.*', 0.8),
            (r'.*malicious.*', 0.9)
        ]
        
        self.benign_patterns = [
            (r'.*benign.*', 0.9),
            (r'.*safe.*', 0.85),
            (r'.*normal.*', 0.8),
            (r'.*appropriate.*', 0.8),
            (r'.*legitimate.*', 0.85),
            (r'.*educational.*', 0.8),
            (r'.*helpful.*', 0.75)
        ]
    
    def classify_ground_truth_with_confidence(self, value: Any, column_name: str = "", 
                                           dataset_context: Dict = None) -> Tuple[Optional[bool], float]:
        """Enhanced classification with confidence scores"""
        if pd.isna(value) or value is None:
            return None, 0.0
        
        str_value = str(value).lower().strip()
        if not str_value:
            return None, 0.0
        
        # Direct keyword matching with weights
        max_malicious_score = 0.0
        max_benign_score = 0.0
        
        for keyword, weight in self.malicious_keywords.items():
            if keyword in str_value or str_value in keyword:
                max_malicious_score = max(max_malicious_score, weight)
        
        for keyword, weight in self.benign_keywords.items():
            if keyword in str_value or str_value in keyword:
                max_benign_score = max(max_benign_score, weight)
        
        # Pattern matching
        for pattern, weight in self.malicious_patterns:
            if re.match(pattern, str_value, re.IGNORECASE):
                max_malicious_score = max(max_malicious_score, weight)
        
        for pattern, weight in self.benign_patterns:
            if re.match(pattern, str_value, re.IGNORECASE):
                max_benign_score = max(max_benign_score, weight)
        
        # Boolean handling with context
        boolean_mappings = {
            'true': True, 'false': False, 'yes': True, 'no': False,
            '1': True, '0': False, 'selected': True, 'not_selected': False
        }
        
        if str_value in boolean_mappings:
            bool_result = boolean_mappings[str_value]
            confidence = 0.8  # High confidence for boolean values
            
            # Context-aware interpretation
            if any(term in column_name.lower() for term in ['jailbreak', 'attack', 'harmful', 'selected']):
                return bool_result, confidence
            elif any(term in column_name.lower() for term in ['benign', 'safe', 'normal']):
                return not bool_result, confidence
            else:
                return bool_result, confidence  # Default: True = malicious
        
        # Determine classification
        if max_malicious_score > max_benign_score and max_malicious_score > 0.5:
            return True, max_malicious_score
        elif max_benign_score > max_malicious_score and max_benign_score > 0.5:
            return False, max_benign_score
        elif max_malicious_score > 0.3 or max_benign_score > 0.3:
            # Lower confidence classification
            if max_malicious_score > max_benign_score:
                return True, max_malicious_score
            else:
                return False, max_benign_score
        
        return None, 0.0
    
    def analyze_ground_truth_distribution(self, df: pd.DataFrame, column_name: str, 
                                        filename: str = "") -> Dict[str, Any]:
        """Comprehensive ground truth analysis with ML readiness"""
        context = {'filename': filename}
        classifications = {}
        confidence_scores = []
        value_counts = Counter()
        
        for value in df[column_name].dropna():
            classification, confidence = self.classify_ground_truth_with_confidence(
                value, column_name, context)
            
            str_val = str(value)
            value_counts[str_val] += 1
            
            if classification is not None:
                classifications[str_val] = {
                    'classification': classification,
                    'confidence': confidence,
                    'count': value_counts[str_val]
                }
                confidence_scores.append(confidence)
        
        # Calculate comprehensive statistics
        total_values = len(df[column_name].dropna())
        classified_values = [v for v in df[column_name].dropna() 
                           if self.classify_ground_truth_with_confidence(v, column_name, context)[0] is not None]
        
        malicious_values = [v for v in df[column_name].dropna() 
                           if self.classify_ground_truth_with_confidence(v, column_name, context)[0] is True]
        
        benign_values = [v for v in df[column_name].dropna() 
                        if self.classify_ground_truth_with_confidence(v, column_name, context)[0] is False]
        
        analysis = {
            'column_name': column_name,
            'total_values': total_values,
            'unique_values': len(df[column_name].dropna().unique()),
            'classified_count': len(classified_values),
            'unclassified_count': total_values - len(classified_values),
            'malicious_count': len(malicious_values),
            'benign_count': len(benign_values),
            'classification_rate': (len(classified_values) / total_values * 100) if total_values > 0 else 0,
            'malicious_rate': (len(malicious_values) / total_values * 100) if total_values > 0 else 0,
            'benign_rate': (len(benign_values) / total_values * 100) if total_values > 0 else 0,
            'avg_confidence': np.mean(confidence_scores) if confidence_scores and NUMPY_AVAILABLE else (sum(confidence_scores) / len(confidence_scores) if confidence_scores else 0),
            'min_confidence': min(confidence_scores) if confidence_scores else 0,
            'max_confidence': max(confidence_scores) if confidence_scores else 0,
            'classifications': classifications,
            'value_distribution': dict(value_counts.most_common()),
            'ml_ready': len(classified_values) > 0 and len(malicious_values) > 0 and len(benign_values) > 0,
            'class_balance': abs(len(malicious_values) - len(benign_values)) / max(1, len(classified_values))
        }
        
        return analysis

class EnhancedPromptResult:
    """Enhanced result class with additional metrics"""
    def __init__(self, prompt="", decision="", latency_ms=0.0, category="", 
                 original_type="", confidence_score=0.0, severity_scores="", 
                 error_message="", timestamp="", ground_truth_binary=None,
                 ground_truth_confidence=0.0, prompt_length=0, prompt_complexity=0.0):
        self.prompt = prompt
        self.decision = decision
        self.latency_ms = latency_ms
        self.category = category
        self.original_type = original_type
        self.confidence_score = confidence_score
        self.severity_scores = severity_scores
        self.error_message = error_message
        self.timestamp = timestamp
        self.ground_truth_binary = ground_truth_binary
        self.ground_truth_confidence = ground_truth_confidence
        self.prompt_length = prompt_length
        self.prompt_complexity = prompt_complexity

class AdvancedMetricsCalculator:
    """Advanced performance metrics with statistical analysis"""
    
    @staticmethod
    def calculate_comprehensive_metrics(results: List[EnhancedPromptResult]) -> Dict[str, Any]:
        """Calculate comprehensive performance metrics"""
        
        if not SKLEARN_AVAILABLE:
            return {"error": "scikit-learn required for advanced metrics"}
        
        # Filter valid results
        valid_results = [r for r in results 
                        if r.decision in ["BLOCKED", "ALLOWED"] and r.ground_truth_binary is not None]
        
        if len(valid_results) < 5:
            return {"error": f"Insufficient valid results: {len(valid_results)}"}
        
        # Prepare data
        y_true = [r.ground_truth_binary for r in valid_results]
        y_pred = [r.decision == 'BLOCKED' for r in valid_results]
        y_scores = [r.confidence_score for r in valid_results]
        latencies = [r.latency_ms for r in valid_results]
        gt_confidences = [r.ground_truth_confidence for r in valid_results]
        
        # Basic metrics
        tn, fp, fn, tp = confusion_matrix(y_true, y_pred).ravel()
        
        accuracy = accuracy_score(y_true, y_pred)
        precision = precision_score(y_true, y_pred, zero_division=0)
        recall = recall_score(y_true, y_pred, zero_division=0)
        f1 = f1_score(y_true, y_pred, zero_division=0)
        
        # Advanced metrics
        try:
            auc_score = roc_auc_score(y_true, y_scores) if len(set(y_true)) > 1 else 0.0
        except:
            auc_score = 0.0
        
        # Rates
        tpr = recall  # True Positive Rate
        tnr = tn / (tn + fp) if (tn + fp) > 0 else 0  # True Negative Rate
        fpr = fp / (fp + tn) if (fp + tn) > 0 else 0  # False Positive Rate
        fnr = fn / (fn + tp) if (fn + tp) > 0 else 0  # False Negative Rate
        
        # Business metrics
        malicious_detection_rate = (tp / (tp + fn) * 100) if (tp + fn) > 0 else 0
        benign_approval_rate = (tn / (tn + fp) * 100) if (tn + fp) > 0 else 0
        
        # Performance statistics
        if NUMPY_AVAILABLE:
            avg_latency = float(np.mean(latencies))
            std_latency = float(np.std(latencies))
            p95_latency = float(np.percentile(latencies, 95))
            avg_confidence = float(np.mean(y_scores))
            avg_gt_confidence = float(np.mean(gt_confidences))
        else:
            avg_latency = sum(latencies) / len(latencies)
            std_latency = 0.0
            p95_latency = max(latencies)
            avg_confidence = sum(y_scores) / len(y_scores)
            avg_gt_confidence = sum(gt_confidences) / len(gt_confidences)
        
        # Classification report
        try:
            class_report = classification_report(y_true, y_pred, output_dict=True, zero_division=0)
        except:
            class_report = {}
        
        # Confidence analysis
        correct_predictions = [i for i in range(len(y_true)) if y_true[i] == y_pred[i]]
        incorrect_predictions = [i for i in range(len(y_true)) if y_true[i] != y_pred[i]]
        
        if NUMPY_AVAILABLE:
            correct_confidence = np.mean([y_scores[i] for i in correct_predictions]) if correct_predictions else 0
            incorrect_confidence = np.mean([y_scores[i] for i in incorrect_predictions]) if incorrect_predictions else 0
        else:
            correct_confidence = sum(y_scores[i] for i in correct_predictions) / len(correct_predictions) if correct_predictions else 0
            incorrect_confidence = sum(y_scores[i] for i in incorrect_predictions) / len(incorrect_predictions) if incorrect_predictions else 0
        
        metrics = {
            # Dataset info
            'total_samples': len(valid_results),
            'malicious_samples': sum(y_true),
            'benign_samples': len(y_true) - sum(y_true),
            'class_balance_ratio': (len(y_true) - sum(y_true)) / max(sum(y_true), 1),
            
            # Confusion Matrix
            'true_positives': int(tp),
            'true_negatives': int(tn),
            'false_positives': int(fp),
            'false_negatives': int(fn),
            
            # Core Metrics
            'accuracy': round(accuracy, 4),
            'precision': round(precision, 4),
            'recall': round(recall, 4),
            'f1_score': round(f1, 4),
            'auc_score': round(auc_score, 4),
            
            # Rates
            'true_positive_rate': round(tpr, 4),
            'true_negative_rate': round(tnr, 4),
            'false_positive_rate': round(fpr, 4),
            'false_negative_rate': round(fnr, 4),
            
            # Business Impact
            'malicious_detection_rate': round(malicious_detection_rate, 2),
            'benign_approval_rate': round(benign_approval_rate, 2),
            
            # Performance
            'avg_latency_ms': round(avg_latency, 2),
            'std_latency_ms': round(std_latency, 2),
            'p95_latency_ms': round(p95_latency, 2),
            'avg_confidence_score': round(avg_confidence, 4),
            'avg_ground_truth_confidence': round(avg_gt_confidence, 4),
            
            # Advanced Analysis
            'correct_prediction_confidence': round(correct_confidence, 4),
            'incorrect_prediction_confidence': round(incorrect_confidence, 4),
            'confidence_discrimination': round(abs(correct_confidence - incorrect_confidence), 4),
            
            # Classification Report
            'classification_report': class_report,
            
            # Quality indicators
            'high_confidence_predictions': sum(1 for score in y_scores if score > 0.8),
            'low_confidence_predictions': sum(1 for score in y_scores if score < 0.3),
            'prediction_confidence_rate': round(sum(1 for score in y_scores if score > 0.5) / len(y_scores) * 100, 2)
        }
        
        return metrics

class EnhancedAzureContentSafetyProcessor:
    """Enhanced processor with advanced features"""
    
    def __init__(self, 
                 endpoint_url: str,
                 api_key: str,
                 max_concurrent_requests: int = 10,
                 rate_limit_per_minute: int = 60,
                 timeout_seconds: int = 30,
                 output_directory: str = "azure_results",
                 prompt_column: str = None,
                 ground_truth_column: str = None,
                 severity_threshold: int = 2,
                 auto_detect_schema: bool = True):
        
        self.endpoint_url = endpoint_url.rstrip('/')
        self.api_key = api_key
        self.max_concurrent_requests = max_concurrent_requests
        self.rate_limit_per_minute = rate_limit_per_minute
        self.timeout_seconds = timeout_seconds
        self.severity_threshold = severity_threshold
        self.auto_detect_schema = auto_detect_schema
        
        self.prompt_column = prompt_column
        self.ground_truth_column = ground_truth_column
        
        self.output_directory = Path(output_directory)
        self.output_directory.mkdir(exist_ok=True)
        
        self.semaphore = asyncio.Semaphore(max_concurrent_requests)
        self.request_times = []
        self.results = []
        
        self.working_endpoint = None
        self.api_version = None
        
        # Initialize components
        self.schema_detector = DatasetSchemaDetector()
        self.gt_mapper = AdvancedGroundTruthMapper()
        self.metrics_calculator = AdvancedMetricsCalculator()
        
        # Statistics
        self.start_time = None
        self.total_requests = 0
        self.successful_requests = 0
        self.failed_requests = 0
        self.dataset_info = {}

    async def discover_working_endpoint(self) -> bool:
        """Enhanced endpoint discovery with better error handling"""
        test_endpoints = [
            (f"{self.endpoint_url}/contentsafety/text:analyze?api-version=2023-10-01", "2023-10-01"),
            (f"{self.endpoint_url}/contentsafety/text:shieldPrompt?api-version=2024-02-15-preview", "2024-02-15-preview"),
            (f"{self.endpoint_url}/contentsafety/text:analyze?api-version=2024-02-15-preview", "2024-02-15-preview"),
            (f"{self.endpoint_url}/contentsafety/text:analyze?api-version=2024-09-01", "2024-09-01"),
        ]
        
        headers = {
            'Content-Type': 'application/json',
            'Ocp-Apim-Subscription-Key': self.api_key,
            'User-Agent': 'Azure-ContentSafety-EnhancedProcessor/5.0'
        }
        
        test_payload = {
            'text': 'Hello world test message for endpoint discovery',
            'categories': ['Hate', 'SelfHarm', 'Sexual', 'Violence'],
            'outputType': 'FourSeverityLevels'
        }
        
        shield_payload = {
            'userPrompt': 'Hello world test message for endpoint discovery',
            'documents': []
        }
        
        connector = aiohttp.TCPConnector(limit=1)
        timeout = aiohttp.ClientTimeout(total=15)
        
        try:
            async with aiohttp.ClientSession(connector=connector, timeout=timeout) as session:
                for endpoint_url, api_version in test_endpoints:
                    try:
                        logger.info(f"Testing endpoint: {endpoint_url}")
                        
                        current_payload = shield_payload if 'shieldPrompt' in endpoint_url else test_payload
                        
                        async with session.post(endpoint_url, headers=headers, json=current_payload) as response:
                            if response.status == 200:
                                result = await response.json()
                                logger.info(f"✅ Working endpoint found: {endpoint_url}")
                                
                                self.working_endpoint = endpoint_url
                                self.api_version = api_version
                                return True
                            else:
                                error_text = await response.text()
                                logger.warning(f"❌ Endpoint failed {response.status}: {error_text[:200]}")
                                
                    except Exception as e:
                        logger.warning(f"❌ Endpoint error: {str(e)}")
                        continue
                        
        except Exception as e:
            logger.error(f"❌ Connection error: {str(e)}")
        
        logger.error("❌ No working endpoint found!")
        return False

    def load_and_analyze_dataset(self, file_path: str) -> List[Dict[str, Any]]:
        """Enhanced dataset loading with comprehensive analysis"""
        file_path = Path(file_path)
        
        if not file_path.exists():
            raise FileNotFoundError(f"Input file not found: {file_path}")
        
        # Load dataset
        try:
            if file_path.suffix.lower() == '.csv':
                encodings = ['utf-8', 'utf-8-sig', 'iso-8859-1', 'cp1252']
                df = None
                
                for encoding in encodings:
                    try:
                        df = pd.read_csv(file_path, encoding=encoding)
                        logger.info(f"Loaded CSV with {encoding} encoding")
                        break
                    except UnicodeDecodeError:
                        continue
                
                if df is None:
                    raise ValueError("Could not read CSV with any encoding")
            else:
                raise ValueError(f"Unsupported file format: {file_path.suffix}")
            
            logger.info(f"Dataset loaded: {len(df)} rows, {len(df.columns)} columns")
            
            # Schema detection
            if self.auto_detect_schema:
                schema_info = self.schema_detector.detect_schema(df, file_path.name)
                logger.info(f"Detected schema: {schema_info['detected_schema']} (confidence: {schema_info['confidence']:.2f})")
                
                if schema_info['schema_info']:
                    logger.info(f"Schema description: {schema_info['schema_info']['description']}")
                
                self.dataset_info = schema_info
            
            # Enhanced column detection
            prompt_col = self._enhanced_detect_prompt_column(df, file_path.name)
            gt_col = self._enhanced_detect_ground_truth_column(df, file_path.name)
            
            # Comprehensive ground truth analysis
            if gt_col:
                gt_analysis = self.gt_mapper.analyze_ground_truth_distribution(df, gt_col, file_path.name)
                logger.info(f"Ground Truth Analysis for '{gt_col}':")
                logger.info(f"  Classification rate: {gt_analysis['classification_rate']:.1f}%")
                logger.info(f"  Malicious: {gt_analysis['malicious_count']} ({gt_analysis['malicious_rate']:.1f}%)")
                logger.info(f"  Benign: {gt_analysis['benign_count']} ({gt_analysis['benign_rate']:.1f}%)")
                logger.info(f"  Average confidence: {gt_analysis['avg_confidence']:.3f}")
                logger.info(f"  ML ready: {gt_analysis['ml_ready']}")
                logger.info(f"  Class balance: {gt_analysis['class_balance']:.3f}")
                
                self.dataset_info['ground_truth_analysis'] = gt_analysis
            
            # Process prompts
            prompts = []
            valid_count = 0
            
            for idx, row in df.iterrows():
                try:
                    prompt_text = str(row[prompt_col]).strip()
                    
                    if pd.isna(row[prompt_col]) or not prompt_text or prompt_text.lower() in ['nan', 'null', '']:
                        continue
                    
                    # Enhanced ground truth processing
                    if gt_col and gt_col in row.index and not pd.isna(row[gt_col]):
                        category = str(row[gt_col]).strip()
                        gt_binary, gt_confidence = self.gt_mapper.classify_ground_truth_with_confidence(
                            row[gt_col], gt_col, {'filename': file_path.name})
                    else:
                        category = 'unknown'
                        gt_binary = None
                        gt_confidence = 0.0
                    
                    # Calculate prompt features
                    prompt_length = len(prompt_text)
                    prompt_complexity = self._calculate_prompt_complexity(prompt_text)
                    
                    prompts.append({
                        'prompt': prompt_text,
                        'category': category,
                        'ground_truth_binary': gt_binary,
                        'ground_truth_confidence': gt_confidence,
                        'prompt_length': prompt_length,
                        'prompt_complexity': prompt_complexity,
                        'row_index': idx
                    })
                    valid_count += 1
                    
                except Exception as e:
                    logger.warning(f"Error processing row {idx}: {str(e)}")
                    continue
            
            logger.info(f"Processed {valid_count} valid prompts from {len(df)} total rows")
            
            # Dataset statistics
            self._log_dataset_statistics(prompts)
            
            return prompts
            
        except Exception as e:
            logger.error(f"Error loading dataset: {str(e)}")
            raise

    def _enhanced_detect_prompt_column(self, df: pd.DataFrame, filename: str = "") -> str:
        """Enhanced prompt column detection with schema awareness"""
        
        # Use schema info if available
        if hasattr(self, 'dataset_info') and self.dataset_info.get('schema_info'):
            schema_prompts = self.dataset_info['schema_info']['prompt_columns']
            for col_candidate in schema_prompts:
                for col in df.columns:
                    if col.lower() == col_candidate.lower():
                        logger.info(f"Schema-based prompt column detection: '{col}'")
                        return col
        
        # Manual override
        if self.prompt_column:
            if self.prompt_column not in df.columns:
                raise ValueError(f"Specified prompt column '{self.prompt_column}' not found")
            return self.prompt_column
        
        # Comprehensive candidate list
        prompt_candidates = [
            'prompt', 'text', 'input', 'query', 'content', 'message',
            'behavior', 'goal', 'target', 'jailbreak_query', 'redteam_query',
            'user_input', 'question', 'request', 'instruction', 'command'
        ]
        
        # Score-based selection
        best_col = None
        best_score = 0
        
        for col in df.columns:
            score = 0
            col_lower = col.lower()
            
            # Exact matches
            if col_lower in prompt_candidates:
                score += 100
            
            # Partial matches
            for candidate in prompt_candidates:
                if candidate in col_lower or col_lower in candidate:
                    score += 50
            
            # Content analysis
            if len(df) > 0:
                sample_text = str(df[col].iloc[0])
                if len(sample_text) > 20:  # Looks like meaningful text
                    score += 30
                if any(word in sample_text.lower() for word in ['how', 'what', 'tell', 'explain', 'write']):
                    score += 20
            
            if score > best_score:
                best_score = score
                best_col = col
        
        if best_col and best_score > 50:
            logger.info(f"Auto-detected prompt column: '{best_col}' (score: {best_score})")
            return best_col
        
        # Fallback to user selection
        logger.error("Could not auto-detect prompt column.")
        logger.error(f"Available columns: {list(df.columns)}")
        self._show_column_samples(df)
        raise ValueError("Could not auto-detect prompt column. Please specify --prompt-column")

    def _enhanced_detect_ground_truth_column(self, df: pd.DataFrame, filename: str = "") -> Optional[str]:
        """Enhanced ground truth detection with confidence scoring"""
        
        # Use schema info if available
        if hasattr(self, 'dataset_info') and self.dataset_info.get('schema_info'):
            schema_gt = self.dataset_info['schema_info']['ground_truth_columns']
            for col_candidate in schema_gt:
                for col in df.columns:
                    if col.lower() == col_candidate.lower():
                        # Verify with analysis
                        analysis = self.gt_mapper.analyze_ground_truth_distribution(df, col, filename)
                        if analysis['classification_rate'] > 30:
                            logger.info(f"Schema-based GT column detection: '{col}' ({analysis['classification_rate']:.1f}% classified)")
                            return col
        
        # Manual override
        if self.ground_truth_column:
            if self.ground_truth_column not in df.columns:
                raise ValueError(f"Specified ground truth column '{self.ground_truth_column}' not found")
            return self.ground_truth_column
        
        # Comprehensive analysis of all potential columns
        gt_candidates = [
            'type', 'category', 'label', 'class', 'ground_truth', 'target',
            'selected_mini', 'transfer_from_llm', 'attack_type', 'safety_label',
            'source', 'from', 'origin', 'is_jailbreak', 'is_harmful', 'is_safe',
            'is_malicious', 'harmful', 'attack', 'violation_type', 'risk_category'
        ]
        
        best_col = None
        best_score = 0
        
        for col in df.columns:
            analysis = self.gt_mapper.analyze_ground_truth_distribution(df, col, filename)
            
            # Calculate composite score
            score = analysis['classification_rate']
            
            # Bonus for having both classes
            if analysis['malicious_count'] > 0 and analysis['benign_count'] > 0:
                score += 25
            
            # Bonus for high confidence
            score += analysis['avg_confidence'] * 20
            
            # Bonus for being in candidate list
            if col.lower() in gt_candidates:
                score += 20
            
            # Penalty for too many unclassified values
            if analysis['unclassified_count'] > analysis['total_values'] * 0.7:
                score -= 30
            
            logger.debug(f"GT column '{col}': score={score:.1f}, classification_rate={analysis['classification_rate']:.1f}%")
            
            if score > best_score:
                best_score = score
                best_col = col
        
        if best_col and best_score > 40:
            logger.info(f"Auto-detected ground truth column: '{best_col}' (score: {best_score:.1f})")
            return best_col
        
        logger.warning("Could not auto-detect reliable ground truth column")
        return None

    def _calculate_prompt_complexity(self, prompt: str) -> float:
        """Calculate prompt complexity score"""
        if not prompt:
            return 0.0
        
        # Basic features
        word_count = len(prompt.split())
        char_count = len(prompt)
        sentence_count = len([s for s in prompt.split('.') if s.strip()])
        
        # Advanced features
        unique_words = len(set(prompt.lower().split()))
        avg_word_length = sum(len(word) for word in prompt.split()) / max(word_count, 1)
        
        # Special characters and patterns
        special_chars = sum(1 for c in prompt if not c.isalnum() and not c.isspace())
        
        # Complexity score (normalized 0-1)
        complexity = (
            min(word_count / 100, 1) * 0.3 +
            min(unique_words / word_count, 1) * 0.2 +
            min(avg_word_length / 10, 1) * 0.2 +
            min(special_chars / char_count, 1) * 0.1 +
            min(sentence_count / 10, 1) * 0.2
        )
        
        return complexity

    def _log_dataset_statistics(self, prompts: List[Dict[str, Any]]):
        """Log comprehensive dataset statistics"""
        if not prompts:
            return
        
        # Basic stats
        total = len(prompts)
        with_gt = sum(1 for p in prompts if p['ground_truth_binary'] is not None)
        malicious = sum(1 for p in prompts if p['ground_truth_binary'] is True)
        benign = sum(1 for p in prompts if p['ground_truth_binary'] is False)
        
        # Prompt characteristics
        if NUMPY_AVAILABLE:
            lengths = [p['prompt_length'] for p in prompts]
            complexities = [p['prompt_complexity'] for p in prompts]
            
            avg_length = np.mean(lengths)
            std_length = np.std(lengths)
            avg_complexity = np.mean(complexities)
        else:
            avg_length = sum(p['prompt_length'] for p in prompts) / len(prompts)
            std_length = 0
            avg_complexity = sum(p['prompt_complexity'] for p in prompts) / len(prompts)
        
        logger.info("DATASET STATISTICS:")
        logger.info(f"  Total prompts: {total:,}")
        logger.info(f"  With ground truth: {with_gt:,} ({with_gt/total*100:.1f}%)")
        logger.info(f"  Malicious: {malicious:,} ({malicious/with_gt*100:.1f}% of labeled)")
        logger.info(f"  Benign: {benign:,} ({benign/with_gt*100:.1f}% of labeled)")
        logger.info(f"  Average prompt length: {avg_length:.1f} ± {std_length:.1f} chars")
        logger.info(f"  Average complexity: {avg_complexity:.3f}")

    def _show_column_samples(self, df: pd.DataFrame):
        """Show sample data for column selection"""
        if len(df) > 0:
            logger.info("Sample data preview:")
            for col in df.columns:
                sample_val = str(df[col].iloc[0])[:100]
                logger.info(f"  {col}: {sample_val}{'...' if len(str(df[col].iloc[0])) > 100 else ''}")

    async def _rate_limit_check(self):
        """Enhanced rate limiting"""
        current_time = time.time()
        self.request_times = [t for t in self.request_times if current_time - t < 60]
        
        if len(self.request_times) >= self.rate_limit_per_minute:
            sleep_time = 60 - (current_time - self.request_times[0])
            if sleep_time > 0:
                logger.info(f"Rate limit reached, sleeping {sleep_time:.2f}s")
                await asyncio.sleep(sleep_time)
                current_time = time.time()
                self.request_times = [t for t in self.request_times if current_time - t < 60]
        
        self.request_times.append(current_time)

    async def _call_content_safety_api(self, session: aiohttp.ClientSession, prompt: str) -> Dict[str, Any]:
        """Enhanced API call with better error handling"""
        if not self.working_endpoint:
            return {'success': False, 'error': 'No working endpoint', 'latency_ms': 0}
        
        headers = {
            'Content-Type': 'application/json',
            'Ocp-Apim-Subscription-Key': self.api_key,
            'User-Agent': 'Azure-ContentSafety-EnhancedProcessor/5.0'
        }
        
        # Prepare payload based on endpoint type
        if 'shieldPrompt' in self.working_endpoint:
            payload = {'userPrompt': prompt, 'documents': []}
        else:
            payload = {
                'text': prompt,
                'categories': ['Hate', 'SelfHarm', 'Sexual', 'Violence'],
                'outputType': 'FourSeverityLevels'
            }
        
        start_time = time.time()
        
        try:
            self.total_requests += 1
            
            async with session.post(
                self.working_endpoint, 
                headers=headers, 
                json=payload,
                timeout=aiohttp.ClientTimeout(total=self.timeout_seconds)
            ) as response:
                latency_ms = (time.time() - start_time) * 1000
                
                if response.status == 200:
                    result = await response.json()
                    self.successful_requests += 1
                    return {'success': True, 'data': result, 'latency_ms': latency_ms}
                else:
                    self.failed_requests += 1
                    error_text = await response.text()
                    return {
                        'success': False, 
                        'error': f"HTTP {response.status}: {error_text[:200]}", 
                        'latency_ms': latency_ms
                    }
                    
        except asyncio.TimeoutError:
            self.failed_requests += 1
            return {'success': False, 'error': 'Request timeout', 'latency_ms': (time.time() - start_time) * 1000}
        except Exception as e:
            self.failed_requests += 1
            return {'success': False, 'error': str(e), 'latency_ms': (time.time() - start_time) * 1000}

    def _parse_content_safety_response(self, response_data: Dict[str, Any]) -> Tuple[str, float, str]:
        """Enhanced response parsing"""
        try:
            # Prompt Shield response
            if 'attackDetected' in response_data:
                attack_detected = response_data.get('attackDetected', False)
                decision = "BLOCKED" if attack_detected else "ALLOWED"
                confidence_score = 0.0
                
                if 'analyzeResults' in response_data:
                    for result in response_data['analyzeResults']:
                        confidence_score = max(confidence_score, result.get('confidenceScore', 0.0))
                
                return decision, confidence_score, f"Prompt Shield: Attack={attack_detected}"
            
            # Content Safety response
            elif 'categoriesAnalysis' in response_data:
                categories_analysis = response_data.get('categoriesAnalysis', [])
                
                if not categories_analysis:
                    return "ALLOWED", 0.0, "No categories analyzed"
                
                severity_details = []
                max_severity = 0
                blocked_categories = []
                total_severity_score = 0
                
                for category in categories_analysis:
                    cat_name = category.get('category', 'Unknown')
                    severity = category.get('severity', 0)
                    
                    severity_details.append(f"{cat_name}:{severity}")
                    max_severity = max(max_severity, severity)
                    total_severity_score += severity
                    
                    if severity > self.severity_threshold:
                        blocked_categories.append(f"{cat_name}({severity})")
                
                decision = "BLOCKED" if max_severity > self.severity_threshold else "ALLOWED"
                
                # Enhanced confidence calculation
                confidence_score = min(max_severity / 6.0, 1.0) if max_severity > 0 else 0.0
                if len(categories_analysis) > 1:
                    # Consider overall severity distribution
                    avg_severity = total_severity_score / len(categories_analysis)
                    confidence_score = (confidence_score + min(avg_severity / 6.0, 1.0)) / 2
                
                severity_summary = "|".join(severity_details)
                if blocked_categories:
                    severity_summary += f" [FLAGGED: {','.join(blocked_categories)}]"
                
                return decision, confidence_score, severity_summary
            
            else:
                return "ALLOWED", 0.0, "Unknown response format"
                
        except Exception as e:
            return "ERROR", 0.0, f"Parse error: {str(e)}"

    async def _process_single_prompt(self, session: aiohttp.ClientSession, prompt_data: Dict[str, Any]):
        """Enhanced single prompt processing"""
        async with self.semaphore:
            await self._rate_limit_check()
            
            prompt = prompt_data['prompt']
            
            # Truncate very long prompts
            if len(prompt) > 10000:
                prompt = prompt[:10000] + "..."
            
            api_result = await self._call_content_safety_api(session, prompt)
            
            if api_result['success']:
                try:
                    decision, confidence_score, severity_scores = self._parse_content_safety_response(api_result['data'])
                    
                    return EnhancedPromptResult(
                        prompt=prompt_data['prompt'],
                        decision=decision,
                        latency_ms=api_result['latency_ms'],
                        category=prompt_data['category'],
                        original_type=prompt_data['category'],
                        confidence_score=confidence_score,
                        severity_scores=severity_scores,
                        error_message="",
                        timestamp=datetime.now().isoformat(),
                        ground_truth_binary=prompt_data.get('ground_truth_binary'),
                        ground_truth_confidence=prompt_data.get('ground_truth_confidence', 0.0),
                        prompt_length=prompt_data.get('prompt_length', 0),
                        prompt_complexity=prompt_data.get('prompt_complexity', 0.0)
                    )
                    
                except Exception as e:
                    return EnhancedPromptResult(
                        prompt=prompt_data['prompt'],
                        decision="ERROR",
                        latency_ms=api_result['latency_ms'],
                        category=prompt_data['category'],
                        original_type=prompt_data['category'],
                        confidence_score=0.0,
                        severity_scores="",
                        error_message=f"Parse error: {str(e)}",
                        timestamp=datetime.now().isoformat(),
                        ground_truth_binary=prompt_data.get('ground_truth_binary'),
                        ground_truth_confidence=prompt_data.get('ground_truth_confidence', 0.0),
                        prompt_length=prompt_data.get('prompt_length', 0),
                        prompt_complexity=prompt_data.get('prompt_complexity', 0.0)
                    )
            else:
                return EnhancedPromptResult(
                    prompt=prompt_data['prompt'],
                    decision="ERROR",
                    latency_ms=api_result['latency_ms'],
                    category=prompt_data['category'],
                    original_type=prompt_data['category'],
                    confidence_score=0.0,
                    severity_scores="",
                    error_message=api_result['error'],
                    timestamp=datetime.now().isoformat(),
                    ground_truth_binary=prompt_data.get('ground_truth_binary'),
                    ground_truth_confidence=prompt_data.get('ground_truth_confidence', 0.0),
                    prompt_length=prompt_data.get('prompt_length', 0),
                    prompt_complexity=prompt_data.get('prompt_complexity', 0.0)
                )

    async def process_batch(self, prompts: List[Dict[str, Any]]) -> List[EnhancedPromptResult]:
        """Enhanced batch processing with progress tracking"""
        
        if not await self.discover_working_endpoint():
            raise ConnectionError("No working endpoint found")
        
        logger.info(f"Processing {len(prompts)} prompts using: {self.working_endpoint}")
        
        self.start_time = time.time()
        
        connector = aiohttp.TCPConnector(
            limit=self.max_concurrent_requests * 2,
            limit_per_host=self.max_concurrent_requests,
            keepalive_timeout=30
        )
        
        try:
            async with aiohttp.ClientSession(connector=connector) as session:
                tasks = [self._process_single_prompt(session, prompt_data) for prompt_data in prompts]
                
                batch_size = min(self.max_concurrent_requests, len(tasks))
                results = []
                
                for i in range(0, len(tasks), batch_size):
                    batch_tasks = tasks[i:i + batch_size]
                    
                    try:
                        batch_results = await asyncio.gather(*batch_tasks, return_exceptions=True)
                        
                        for j, result in enumerate(batch_results):
                            if isinstance(result, Exception):
                                logger.error(f"Task failed: {str(result)}")
                                error_result = EnhancedPromptResult(
                                    prompt=prompts[i + j]['prompt'] if i + j < len(prompts) else "ERROR",
                                    decision="ERROR",
                                    latency_ms=0.0,
                                    category="error",
                                    original_type="error",
                                    confidence_score=0.0,
                                    severity_scores="",
                                    error_message=f"Task exception: {str(result)}",
                                    timestamp=datetime.now().isoformat()
                                )
                                results.append(error_result)
                            else:
                                results.append(result)
                        
                        # Enhanced progress logging
                        processed = min(i + batch_size, len(tasks))
                        elapsed = time.time() - self.start_time
                        rate = processed / elapsed if elapsed > 0 else 0
                        eta = (len(tasks) - processed) / rate if rate > 0 else 0
                        
                        logger.info(f"Progress: {processed}/{len(tasks)} ({processed/len(tasks)*100:.1f}%) | "
                                   f"Rate: {rate:.1f}/sec | Success: {self.successful_requests/self.total_requests*100:.1f}% | "
                                   f"ETA: {eta/60:.1f}min")
                        
                    except Exception as e:
                        logger.error(f"Batch error: {str(e)}")
                        continue
                    
                    if i + batch_size < len(tasks):
                        await asyncio.sleep(0.1)
                        
        except Exception as e:
            logger.error(f"Session error: {str(e)}")
            raise
        
        self.results = results
        
        total_time = time.time() - self.start_time
        logger.info(f"Completed in {total_time/60:.1f} minutes. Success rate: {self.successful_requests/self.total_requests*100:.1f}%")
        
        return results

    def save_comprehensive_results(self, output_file: str = None, input_filename: str = None) -> Dict[str, str]:
        """Save comprehensive results with multiple output formats"""
        if not self.results:
            raise ValueError("No results to save")
        
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        if input_filename:
            input_base = Path(input_filename).stem.replace(' ', '_')
            base_name = f"azure_{input_base}_{timestamp}"
        else:
            base_name = f"azure_content_safety_results_{timestamp}"
        
        output_files = {}
        
        # 1. Main results CSV
        results_file = self.output_directory / f"{base_name}_results.csv"
        self._save_results_csv(results_file)
        output_files['results'] = str(results_file)
        
        # 2. Performance metrics CSV
        metrics_file = self.output_directory / f"{base_name}_metrics.csv"
        self._save_performance_metrics_csv(metrics_file, input_filename)
        output_files['metrics'] = str(metrics_file)
        
        # 3. Dataset analysis JSON
        analysis_file = self.output_directory / f"{base_name}_analysis.json"
        self._save_dataset_analysis_json(analysis_file)
        output_files['analysis'] = str(analysis_file)
        
        # 4. Summary report
        summary_file = self.output_directory / f"{base_name}_summary.txt"
        self._save_summary_report(summary_file, input_filename)
        output_files['summary'] = str(summary_file)
        
        # 5. Excel workbook with all sheets
        if OPENPYXL_AVAILABLE:
            excel_file = self.output_directory / f"{base_name}_complete_analysis.xlsx"
            self._save_excel_workbook(excel_file, input_filename)
            output_files['excel'] = str(excel_file)
        
        logger.info(f"Comprehensive results saved:")
        for file_type, file_path in output_files.items():
            logger.info(f"  {file_type.title()}: {file_path}")
        
        return output_files

    def _save_results_csv(self, output_path: Path):
        """Save enhanced results CSV"""
        data = []
        for result in self.results:
            data.append({
                'prompt': result.prompt,
                'original_type': result.original_type,
                'ground_truth_binary': result.ground_truth_binary,
                'ground_truth_confidence': result.ground_truth_confidence,
                'decision': result.decision,
                'confidence_score': result.confidence_score,
                'severity_scores': result.severity_scores,
                'latency_ms': result.latency_ms,
                'prompt_length': result.prompt_length,
                'prompt_complexity': result.prompt_complexity,
                'error_message': result.error_message,
                'timestamp': result.timestamp
            })
        
        df = pd.DataFrame(data)
        df.to_csv(output_path, index=False, encoding='utf-8')

    def _save_performance_metrics_csv(self, metrics_path: Path, input_filename: str = None):
        """Save comprehensive performance metrics"""
        try:
            metrics = self.metrics_calculator.calculate_comprehensive_metrics(self.results)
            
            if 'error' not in metrics:
                # Create enhanced metrics data
                metrics_data = [
                    ['Metric', 'Value', 'Description', 'Category'],
                    
                    # Dataset Information
                    ['Input File', input_filename or 'Unknown', 'Source dataset', 'Dataset'],
                    ['Processing Date', datetime.now().strftime('%Y-%m-%d %H:%M:%S'), 'Analysis timestamp', 'Dataset'],
                    ['Detected Schema', self.dataset_info.get('detected_schema', 'unknown'), 'Auto-detected schema type', 'Dataset'],
                    ['Schema Confidence', f"{self.dataset_info.get('confidence', 0):.2f}", 'Schema detection confidence', 'Dataset'],
                    ['Total Samples', metrics['total_samples'], 'Valid samples with ground truth', 'Dataset'],
                    ['Malicious Samples', metrics['malicious_samples'], 'Samples labeled as malicious', 'Dataset'],
                    ['Benign Samples', metrics['benign_samples'], 'Samples labeled as benign', 'Dataset'],
                    ['Class Balance Ratio', f"{metrics['class_balance_ratio']:.3f}", 'Benign/Malicious ratio', 'Dataset'],
                    
                    # API Configuration
                    ['API Endpoint', self.working_endpoint or 'Unknown', 'Azure API endpoint used', 'API'],
                    ['API Version', self.api_version or 'Unknown', 'Azure API version', 'API'],
                    ['Severity Threshold', self.severity_threshold, 'Blocking threshold (0-6)', 'API'],
                    ['Concurrent Requests', self.max_concurrent_requests, 'Max parallel requests', 'API'],
                    ['Rate Limit', self.rate_limit_per_minute, 'Requests per minute limit', 'API'],
                    
                    # Confusion Matrix
                    ['True Positives (TP)', metrics['true_positives'], 'Malicious correctly blocked', 'Performance'],
                    ['True Negatives (TN)', metrics['true_negatives'], 'Benign correctly allowed', 'Performance'],
                    ['False Positives (FP)', metrics['false_positives'], 'Benign incorrectly blocked', 'Performance'],
                    ['False Negatives (FN)', metrics['false_negatives'], 'Malicious incorrectly allowed', 'Performance'],
                    
                    # Core Metrics
                    ['Accuracy', metrics['accuracy'], 'Overall prediction accuracy', 'Performance'],
                    ['Precision', metrics['precision'], 'Precision of blocking decisions', 'Performance'],
                    ['Recall (Sensitivity)', metrics['recall'], 'Malicious detection rate', 'Performance'],
                    ['F1 Score', metrics['f1_score'], 'Harmonic mean of precision/recall', 'Performance'],
                    ['AUC Score', metrics['auc_score'], 'Area under ROC curve', 'Performance'],
                    
                    # Rates
                    ['True Positive Rate', metrics['true_positive_rate'], 'Sensitivity/Recall', 'Rates'],
                    ['True Negative Rate', metrics['true_negative_rate'], 'Specificity', 'Rates'],
                    ['False Positive Rate', metrics['false_positive_rate'], 'Type I error rate', 'Rates'],
                    ['False Negative Rate', metrics['false_negative_rate'], 'Type II error rate', 'Rates'],
                    
                    # Business Impact
                    ['Malicious Detection Rate (%)', metrics['malicious_detection_rate'], 'Malicious prompts blocked', 'Business'],
                    ['Benign Approval Rate (%)', metrics['benign_approval_rate'], 'Benign prompts allowed', 'Business'],
                    
                    # Performance Statistics
                    ['Average Latency (ms)', metrics['avg_latency_ms'], 'Mean response time', 'Performance'],
                    ['Std Latency (ms)', metrics['std_latency_ms'], 'Latency standard deviation', 'Performance'],
                    ['95th Percentile Latency (ms)', metrics['p95_latency_ms'], '95% of requests faster than', 'Performance'],
                    ['Average Confidence Score', metrics['avg_confidence_score'], 'Mean prediction confidence', 'Confidence'],
                    ['Average GT Confidence', metrics['avg_ground_truth_confidence'], 'Mean ground truth confidence', 'Confidence'],
                    ['Correct Prediction Confidence', metrics['correct_prediction_confidence'], 'Confidence when correct', 'Confidence'],
                    ['Incorrect Prediction Confidence', metrics['incorrect_prediction_confidence'], 'Confidence when incorrect', 'Confidence'],
                    ['Confidence Discrimination', metrics['confidence_discrimination'], 'Confidence difference (correct vs incorrect)', 'Confidence'],
                    
                    # Quality Indicators
                    ['High Confidence Predictions', metrics['high_confidence_predictions'], 'Predictions with confidence > 0.8', 'Quality'],
                    ['Low Confidence Predictions', metrics['low_confidence_predictions'], 'Predictions with confidence < 0.3', 'Quality'],
                    ['Prediction Confidence Rate (%)', metrics['prediction_confidence_rate'], 'Predictions with confidence > 0.5', 'Quality'],
                    
                    # API Statistics
                    ['Total API Requests', self.total_requests, 'Total requests made', 'API'],
                    ['Successful Requests', self.successful_requests, 'Successful API calls', 'API'],
                    ['Failed Requests', self.failed_requests, 'Failed API calls', 'API'],
                    ['API Success Rate (%)', (self.successful_requests / self.total_requests * 100) if self.total_requests > 0 else 0, 'API reliability', 'API'],
                ]
                
                metrics_df = pd.DataFrame(metrics_data[1:], columns=metrics_data[0])
                metrics_df.to_csv(metrics_path, index=False, encoding='utf-8')
                
                logger.info(f"Performance metrics saved to {metrics_path}")
                
                # Print key metrics
                self._print_key_metrics(metrics)
                
            else:
                logger.error(f"Could not calculate performance metrics: {metrics.get('error')}")
                
        except Exception as e:
            logger.error(f"Error saving performance metrics: {str(e)}")

    def _save_dataset_analysis_json(self, analysis_path: Path):
        """Save comprehensive dataset analysis as JSON"""
        try:
            analysis_data = {
                'dataset_info': self.dataset_info,
                'processing_stats': {
                    'total_requests': self.total_requests,
                    'successful_requests': self.successful_requests,
                    'failed_requests': self.failed_requests,
                    'processing_time_minutes': (time.time() - self.start_time) / 60 if self.start_time else 0,
                    'average_latency_ms': sum(r.latency_ms for r in self.results) / len(self.results) if self.results else 0
                },
                'api_config': {
                    'endpoint': self.working_endpoint,
                    'api_version': self.api_version,
                    'severity_threshold': self.severity_threshold,
                    'max_concurrent_requests': self.max_concurrent_requests,
                    'rate_limit_per_minute': self.rate_limit_per_minute
                },
                'results_summary': {
                    'total_results': len(self.results),
                    'decisions': {
                        'blocked': sum(1 for r in self.results if r.decision == 'BLOCKED'),
                        'allowed': sum(1 for r in self.results if r.decision == 'ALLOWED'),
                        'error': sum(1 for r in self.results if r.decision == 'ERROR')
                    },
                    'ground_truth_distribution': {
                        'malicious': sum(1 for r in self.results if r.ground_truth_binary is True),
                        'benign': sum(1 for r in self.results if r.ground_truth_binary is False),
                        'unknown': sum(1 for r in self.results if r.ground_truth_binary is None)
                    }
                }
            }
            
            with open(analysis_path, 'w', encoding='utf-8') as f:
                json.dump(analysis_data, f, indent=2, default=str)
            
            logger.info(f"Dataset analysis saved to {analysis_path}")
            
        except Exception as e:
            logger.error(f"Error saving dataset analysis: {str(e)}")

    def _save_summary_report(self, summary_path: Path, input_filename: str = None):
        """Save human-readable summary report"""
        try:
            metrics = self.metrics_calculator.calculate_comprehensive_metrics(self.results)
            
            with open(summary_path, 'w', encoding='utf-8') as f:
                f.write("="*80 + "\n")
                f.write("AZURE CONTENT SAFETY - COMPREHENSIVE ANALYSIS REPORT\n")
                f.write("="*80 + "\n\n")
                
                # Dataset Information
                f.write("DATASET INFORMATION\n")
                f.write("-" * 50 + "\n")
                f.write(f"Input File: {input_filename or 'Unknown'}\n")
                f.write(f"Processing Date: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n")
                f.write(f"Detected Schema: {self.dataset_info.get('detected_schema', 'unknown')}\n")
                f.write(f"Schema Confidence: {self.dataset_info.get('confidence', 0):.2f}\n")
                if self.dataset_info.get('schema_info'):
                    f.write(f"Schema Description: {self.dataset_info['schema_info']['description']}\n")
                f.write(f"Total Samples: {len(self.results):,}\n")
                
                if 'error' not in metrics:
                    f.write(f"Valid Samples (with GT): {metrics['total_samples']:,}\n")
                    f.write(f"Malicious Samples: {metrics['malicious_samples']:,}\n")
                    f.write(f"Benign Samples: {metrics['benign_samples']:,}\n")
                    f.write(f"Class Balance: {metrics['class_balance_ratio']:.3f}\n")
                
                f.write("\n")
                
                # API Configuration
                f.write("API CONFIGURATION\n")
                f.write("-" * 50 + "\n")
                f.write(f"Endpoint: {self.working_endpoint or 'Unknown'}\n")
                f.write(f"API Version: {self.api_version or 'Unknown'}\n")
                f.write(f"Severity Threshold: {self.severity_threshold}\n")
                f.write(f"Max Concurrent Requests: {self.max_concurrent_requests}\n")
                f.write(f"Rate Limit: {self.rate_limit_per_minute}/minute\n")
                f.write("\n")
                
                # Performance Results
                if 'error' not in metrics:
                    f.write("PERFORMANCE RESULTS\n")
                    f.write("-" * 50 + "\n")
                    f.write(f"Overall Accuracy: {metrics['accuracy']*100:.2f}%\n")
                    f.write(f"F1 Score: {metrics['f1_score']:.4f}\n")
                    f.write(f"Precision: {metrics['precision']:.4f}\n")
                    f.write(f"Recall: {metrics['recall']:.4f}\n")
                    f.write(f"AUC Score: {metrics['auc_score']:.4f}\n")
                    f.write("\n")
                    
                    f.write("CONFUSION MATRIX\n")
                    f.write("-" * 50 + "\n")
                    f.write(f"True Positives (Malicious Blocked): {metrics['true_positives']:,}\n")
                    f.write(f"True Negatives (Benign Allowed): {metrics['true_negatives']:,}\n")
                    f.write(f"False Positives (Benign Blocked): {metrics['false_positives']:,}\n")
                    f.write(f"False Negatives (Malicious Allowed): {metrics['false_negatives']:,}\n")
                    f.write("\n")
                    
                    f.write("BUSINESS IMPACT\n")
                    f.write("-" * 50 + "\n")
                    f.write(f"Malicious Detection Rate: {metrics['malicious_detection_rate']:.2f}%\n")
                    f.write(f"Benign Approval Rate: {metrics['benign_approval_rate']:.2f}%\n")
                    f.write(f"False Positive Rate: {metrics['false_positive_rate']*100:.2f}%\n")
                    f.write(f"False Negative Rate: {metrics['false_negative_rate']*100:.2f}%\n")
                    f.write("\n")
                    
                    f.write("PERFORMANCE STATISTICS\n")
                    f.write("-" * 50 + "\n")
                    f.write(f"Average Latency: {metrics['avg_latency_ms']:.2f} ms\n")
                    f.write(f"95th Percentile Latency: {metrics['p95_latency_ms']:.2f} ms\n")
                    f.write(f"Average Confidence: {metrics['avg_confidence_score']:.4f}\n")
                    f.write(f"High Confidence Predictions: {metrics['high_confidence_predictions']:,}\n")
                    f.write(f"Low Confidence Predictions: {metrics['low_confidence_predictions']:,}\n")
                    f.write("\n")
                
                # API Statistics
                f.write("API STATISTICS\n")
                f.write("-" * 50 + "\n")
                f.write(f"Total Requests: {self.total_requests:,}\n")
                f.write(f"Successful Requests: {self.successful_requests:,}\n")
                f.write(f"Failed Requests: {self.failed_requests:,}\n")
                f.write(f"Success Rate: {(self.successful_requests/self.total_requests*100):.2f}%\n")
                if self.start_time:
                    total_time = time.time() - self.start_time
                    f.write(f"Total Processing Time: {total_time/60:.2f} minutes\n")
                    f.write(f"Processing Rate: {len(self.results)/(total_time/60):.1f} prompts/minute\n")
                
                f.write("\n" + "="*80 + "\n")
                f.write("Report generated by Azure Content Safety Enhanced Processor v5.0\n")
                f.write("="*80 + "\n")
            
            logger.info(f"Summary report saved to {summary_path}")
            
        except Exception as e:
            logger.error(f"Error saving summary report: {str(e)}")

    def _print_key_metrics(self, metrics: Dict[str, Any]):
        """Print key metrics to console"""
        logger.info("="*60)
        logger.info("KEY PERFORMANCE METRICS")
        logger.info("="*60)
        logger.info(f"Accuracy: {metrics['accuracy']*100:.2f}%")
        logger.info(f"F1 Score: {metrics['f1_score']:.4f}")
        logger.info(f"Precision: {metrics['precision']:.4f}")
        logger.info(f"Recall: {metrics['recall']:.4f}")
        logger.info(f"AUC Score: {metrics['auc_score']:.4f}")
        logger.info("")
        logger.info("CONFUSION MATRIX:")
        logger.info(f"  TP (Malicious Blocked): {metrics['true_positives']:,}")
        logger.info(f"  TN (Benign Allowed): {metrics['true_negatives']:,}")
        logger.info(f"  FP (Benign Blocked): {metrics['false_positives']:,}")
        logger.info(f"  FN (Malicious Allowed): {metrics['false_negatives']:,}")
        logger.info("")
        logger.info("BUSINESS IMPACT:")
        logger.info(f"  Malicious Detection: {metrics['malicious_detection_rate']:.2f}%")
        logger.info(f"  Benign Approval: {metrics['benign_approval_rate']:.2f}%")
        logger.info(f"  Average Latency: {metrics['avg_latency_ms']:.2f} ms")
        logger.info("="*60)

    def _save_excel_workbook(self, excel_path: Path, input_filename: str = None):
        """Save comprehensive Excel workbook with all analysis sheets"""
        if not OPENPYXL_AVAILABLE:
            logger.warning("openpyxl not available - skipping Excel export")
            return
        
        try:
            wb = Workbook()
            
            # Style definitions
            header_fill = PatternFill(start_color="CCE5FF", end_color="CCE5FF", fill_type="solid")
            header_font = Font(bold=True)
            
            # Remove default sheet
            wb.remove(wb.active)
            
            # 1. Summary Sheet
            summary_sheet = wb.create_sheet("Summary")
            self._create_summary_sheet(summary_sheet, header_fill, header_font, input_filename)
            
            # 2. Performance Metrics Sheet
            metrics_sheet = wb.create_sheet("Performance Metrics")
            self._create_metrics_sheet(metrics_sheet, header_fill, header_font, input_filename)
            
            # 3. Detailed Results Sheet
            detailed_sheet = wb.create_sheet("Detailed Results")
            self._create_detailed_results_sheet(detailed_sheet, header_fill, header_font)
            
            # 4. Confusion Matrix & Stats Sheet
            confusion_sheet = wb.create_sheet("Confusion Matrix")
            self._create_confusion_matrix_sheet(confusion_sheet, header_fill, header_font)
            
            # Set summary as active sheet
            wb.active = summary_sheet
            
            wb.save(excel_path)
            logger.info(f"Excel workbook saved to {excel_path}")
            
        except Exception as e:
            logger.error(f"Error saving Excel workbook: {e}")

    def _create_summary_sheet(self, sheet, header_fill, header_font, input_filename: str = None):
        """Create comprehensive summary sheet"""
        try:
            # Header
            sheet.append(["Azure Content Safety Analysis - Summary Report"])
            sheet.append([])
            
            # Dataset Information
            sheet.append(["Dataset Information", "Value"])
            sheet.append(["Input File", input_filename or "Unknown"])
            sheet.append(["Processing Date", datetime.now().strftime('%Y-%m-%d %H:%M:%S')])
            sheet.append(["Detected Schema", self.dataset_info.get('detected_schema', 'unknown')])
            sheet.append(["Schema Confidence", f"{self.dataset_info.get('confidence', 0):.2f}"])
            sheet.append(["Total Prompts Processed", len(self.results)])
            
            # Filter valid results for metrics
            valid_results = [r for r in self.results 
                           if r.decision in ["BLOCKED", "ALLOWED"] and r.ground_truth_binary is not None]
            
            if valid_results:
                malicious_count = sum(1 for r in valid_results if r.ground_truth_binary is True)
                benign_count = sum(1 for r in valid_results if r.ground_truth_binary is False)
                
                sheet.append(["Valid Results (with Ground Truth)", len(valid_results)])
                sheet.append(["Malicious Samples", malicious_count])
                sheet.append(["Benign Samples", benign_count])
                sheet.append(["Class Balance Ratio", f"{benign_count/max(malicious_count, 1):.3f}"])
            
            sheet.append([])
            
            # API Configuration
            sheet.append(["API Configuration", "Value"])
            sheet.append(["Azure Endpoint", self.working_endpoint or "Unknown"])
            sheet.append(["API Version", self.api_version or "Unknown"])
            sheet.append(["Severity Threshold", self.severity_threshold])
            sheet.append(["Max Concurrent Requests", self.max_concurrent_requests])
            sheet.append(["Rate Limit (per minute)", self.rate_limit_per_minute])
            sheet.append([])
            
            # Processing Statistics
            sheet.append(["Processing Statistics", "Value"])
            sheet.append(["Total API Requests", self.total_requests])
            sheet.append(["Successful Requests", self.successful_requests])
            sheet.append(["Failed Requests", self.failed_requests])
            sheet.append(["API Success Rate (%)", f"{(self.successful_requests/self.total_requests*100):.1f}" if self.total_requests > 0 else "0"])
            
            if self.start_time:
                total_time = time.time() - self.start_time
                sheet.append(["Total Processing Time (min)", f"{total_time/60:.1f}"])
                sheet.append(["Processing Rate (prompts/min)", f"{len(self.results)/(total_time/60):.1f}"])
            
            # Average latency
            if self.results:
                avg_latency = sum(r.latency_ms for r in self.results) / len(self.results)
                sheet.append(["Average Latency (ms)", f"{avg_latency:.2f}"])
            
            sheet.append([])
            
            # Decision Distribution
            sheet.append(["Decision Distribution", "Count", "Percentage"])
            decisions = {}
            for result in self.results:
                decisions[result.decision] = decisions.get(result.decision, 0) + 1
            
            for decision, count in decisions.items():
                percentage = (count / len(self.results) * 100) if self.results else 0
                sheet.append([decision, count, f"{percentage:.1f}%"])
            
            # Style headers
            for row in sheet.iter_rows(min_row=1, max_row=sheet.max_row, min_col=1, max_col=1):
                for cell in row:
                    if cell.value and isinstance(cell.value, str) and any(x in cell.value for x in ["Information", "Configuration", "Statistics", "Distribution"]):
                        cell.fill = header_fill
                        cell.font = header_font
                        
        except Exception as e:
            logger.error(f"Error creating summary sheet: {e}")

    def _create_metrics_sheet(self, sheet, header_fill, header_font, input_filename: str = None):
        """Create performance metrics sheet"""
        try:
            metrics = self.metrics_calculator.calculate_comprehensive_metrics(self.results)
            
            if 'error' in metrics:
                sheet.append(["Error calculating metrics", metrics['error']])
                return
            
            # Header
            sheet.append(["Performance Metrics Analysis"])
            sheet.append([])
            sheet.append(["Metric", "Value", "Description", "Category"])
            
            # Core Performance Metrics
            metrics_data = [
                ['Total Samples', metrics['total_samples'], 'Valid samples with ground truth', 'Dataset'],
                ['Malicious Samples', metrics['malicious_samples'], 'Samples labeled as malicious', 'Dataset'],
                ['Benign Samples', metrics['benign_samples'], 'Samples labeled as benign', 'Dataset'],
                ['Class Balance Ratio', f"{metrics['class_balance_ratio']:.3f}", 'Benign/Malicious ratio', 'Dataset'],
                [],
                ['Accuracy', f"{metrics['accuracy']:.4f}", 'Overall prediction accuracy', 'Performance'],
                ['Precision', f"{metrics['precision']:.4f}", 'Precision of blocking decisions', 'Performance'],
                ['Recall (Sensitivity)', f"{metrics['recall']:.4f}", 'Malicious detection rate', 'Performance'],
                ['F1 Score', f"{metrics['f1_score']:.4f}", 'Harmonic mean of precision/recall', 'Performance'],
                ['AUC Score', f"{metrics['auc_score']:.4f}", 'Area under ROC curve', 'Performance'],
                [],
                ['True Positives (TP)', metrics['true_positives'], 'Malicious correctly blocked', 'Confusion Matrix'],
                ['True Negatives (TN)', metrics['true_negatives'], 'Benign correctly allowed', 'Confusion Matrix'],
                ['False Positives (FP)', metrics['false_positives'], 'Benign incorrectly blocked', 'Confusion Matrix'],
                ['False Negatives (FN)', metrics['false_negatives'], 'Malicious incorrectly allowed', 'Confusion Matrix'],
                [],
                ['True Positive Rate', f"{metrics['true_positive_rate']:.4f}", 'Sensitivity/Recall', 'Rates'],
                ['True Negative Rate', f"{metrics['true_negative_rate']:.4f}", 'Specificity', 'Rates'],
                ['False Positive Rate', f"{metrics['false_positive_rate']:.4f}", 'Type I error rate', 'Rates'],
                ['False Negative Rate', f"{metrics['false_negative_rate']:.4f}", 'Type II error rate', 'Rates'],
                [],
                ['Malicious Detection Rate (%)', f"{metrics['malicious_detection_rate']:.2f}", 'Malicious prompts blocked', 'Business Impact'],
                ['Benign Approval Rate (%)', f"{metrics['benign_approval_rate']:.2f}", 'Benign prompts allowed', 'Business Impact'],
                [],
                ['Average Latency (ms)', f"{metrics['avg_latency_ms']:.2f}", 'Mean response time', 'Performance'],
                ['95th Percentile Latency (ms)', f"{metrics['p95_latency_ms']:.2f}", '95% of requests faster than', 'Performance'],
                ['Average Confidence Score', f"{metrics['avg_confidence_score']:.4f}", 'Mean prediction confidence', 'Confidence'],
            ]
            
            for row_data in metrics_data:
                if row_data:  # Skip empty rows
                    sheet.append(row_data)
                else:
                    sheet.append([])
            
            # Style header
            for cell in sheet["3:3"]:
                cell.fill = header_fill
                cell.font = header_font
                
        except Exception as e:
            logger.error(f"Error creating metrics sheet: {e}")

    def _create_detailed_results_sheet(self, sheet, header_fill, header_font):
        """Create detailed results sheet with all prompts"""
        try:
            # Header
            header_row = ["Prompt #", "Prompt Text", "Ground Truth", "Azure Decision", 
                         "Confidence Score", "Latency (ms)", "Severity Scores", "Error Message"]
            sheet.append(header_row)
            
            # Style header
            for cell in sheet["1:1"]:
                cell.fill = header_fill
                cell.font = header_font
            
            # Data rows
            for i, result in enumerate(self.results, 1):
                prompt_text = result.prompt
                # Truncate long prompts for readability
                if len(prompt_text) > 100:
                    prompt_text = prompt_text[:97] + '...'
                
                # Ground truth
                ground_truth_value = "N/A"
                if result.ground_truth_binary is not None:
                    ground_truth_value = "Threat" if result.ground_truth_binary else "Safe"
                
                row = [
                    i,  # Prompt number
                    prompt_text,  # Prompt text
                    ground_truth_value,  # Ground truth
                    result.decision,  # Azure decision
                    f"{result.confidence_score:.4f}" if result.confidence_score else "N/A",  # Confidence
                    f"{result.latency_ms:.2f}",  # Latency
                    result.severity_scores or "N/A",  # Severity scores
                    result.error_message or "None"  # Error message
                ]
                
                sheet.append(row)
            
            # Auto-adjust column widths
            for column in sheet.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)  # Cap at 50 characters
                sheet.column_dimensions[column_letter].width = adjusted_width
                
        except Exception as e:
            logger.error(f"Error creating detailed results sheet: {e}")

    def _create_confusion_matrix_sheet(self, sheet, header_fill, header_font):
        """Create confusion matrix and additional statistics sheet"""
        try:
            metrics = self.metrics_calculator.calculate_comprehensive_metrics(self.results)
            
            if 'error' in metrics:
                sheet.append(["Error calculating metrics", metrics['error']])
                return
            
            # Confusion Matrix
            sheet.append(["Confusion Matrix Analysis"])
            sheet.append([])
            sheet.append(["", "Predicted Malicious", "Predicted Benign", "Total"])
            sheet.append(["Actual Malicious", metrics['true_positives'], metrics['false_negatives'], 
                         metrics['true_positives'] + metrics['false_negatives']])
            sheet.append(["Actual Benign", metrics['false_positives'], metrics['true_negatives'], 
                         metrics['false_positives'] + metrics['true_negatives']])
            sheet.append(["Total", 
                         metrics['true_positives'] + metrics['false_positives'],
                         metrics['false_negatives'] + metrics['true_negatives'],
                         metrics['total_samples']])
            sheet.append([])
            
            # Business Impact Analysis
            sheet.append(["Business Impact Analysis"])
            sheet.append([])
            sheet.append(["Metric", "Value", "Impact"])
            sheet.append(["Malicious Blocked (TP)", metrics['true_positives'], "✅ Correctly protected"])
            sheet.append(["Benign Allowed (TN)", metrics['true_negatives'], "✅ Good user experience"])
            sheet.append(["Benign Blocked (FP)", metrics['false_positives'], "❌ Poor user experience"])
            sheet.append(["Malicious Allowed (FN)", metrics['false_negatives'], "⚠️ Security risk"])
            sheet.append([])
            
            # Quality Indicators
            sheet.append(["Quality Indicators"])
            sheet.append([])
            sheet.append(["Indicator", "Value", "Assessment"])
            
            # Assessment logic
            accuracy_assessment = "Excellent" if metrics['accuracy'] > 0.9 else "Good" if metrics['accuracy'] > 0.8 else "Needs Improvement"
            precision_assessment = "Excellent" if metrics['precision'] > 0.9 else "Good" if metrics['precision'] > 0.8 else "Needs Improvement"
            recall_assessment = "Excellent" if metrics['recall'] > 0.9 else "Good" if metrics['recall'] > 0.8 else "Needs Improvement"
            
            sheet.append([f"Accuracy ({metrics['accuracy']:.1%})", f"{metrics['accuracy']:.4f}", accuracy_assessment])
            sheet.append([f"Precision ({metrics['precision']:.1%})", f"{metrics['precision']:.4f}", precision_assessment])
            sheet.append([f"Recall ({metrics['recall']:.1%})", f"{metrics['recall']:.4f}", recall_assessment])
            sheet.append([f"F1 Score", f"{metrics['f1_score']:.4f}", ""])
            
            # Style headers
            for row_num in [1, 8, 16]:
                try:
                    cell = sheet.cell(row=row_num, column=1)
                    cell.fill = header_fill
                    cell.font = header_font
                except:
                    pass
                    
        except Exception as e:
            logger.error(f"Error creating confusion matrix sheet: {e}")

    def _handle_list_columns(self, input_file: str):
        """Handle --list-columns option"""
        try:
            file_path = Path(input_file)
            if file_path.suffix.lower() == '.csv':
                df = pd.read_csv(file_path, nrows=5)  # Read small sample
                
                print("\n" + "="*80)
                print(f"COLUMN ANALYSIS FOR: {file_path.name}")
                print("="*80)
                
                # Schema detection
                schema_info = self.schema_detector.detect_schema(df, file_path.name)
                print(f"Detected Schema: {schema_info['detected_schema']} (confidence: {schema_info['confidence']:.2f})")
                if schema_info['schema_info']:
                    print(f"Description: {schema_info['schema_info']['description']}")
                print()
                
                # Column list
                print("AVAILABLE COLUMNS:")
                print("-" * 40)
                for i, col in enumerate(df.columns, 1):
                    print(f"  {i:2d}. {col}")
                
                # Sample data
                print("\nSAMPLE DATA:")
                print("-" * 40)
                for col in df.columns[:5]:  # Show first 5 columns
                    sample_val = str(df[col].iloc[0])[:60]
                    print(f"  {col}: {sample_val}{'...' if len(str(df[col].iloc[0])) > 60 else ''}")
                
                # Ground truth analysis
                print("\nGROUND TRUTH ANALYSIS:")
                print("-" * 40)
                for col in df.columns:
                    analysis = self.gt_mapper.analyze_ground_truth_distribution(df, col, file_path.name)
                    if analysis['classification_rate'] > 10:
                        print(f"  {col}: {analysis['classification_rate']:.1f}% classified "
                              f"({analysis['malicious_count']} malicious, {analysis['benign_count']} benign)")
                
                print("\n" + "="*80)
                
        except Exception as e:
            logger.error(f"Error analyzing columns: {str(e)}")

    def _print_dataset_analysis_summary(self, prompts: List[Dict[str, Any]]):
        """Print comprehensive dataset analysis summary"""
        print("\n" + "="*80)
        print("DATASET ANALYSIS SUMMARY")
        print("="*80)
        
        # Basic statistics
        total = len(prompts)
        with_gt = sum(1 for p in prompts if p['ground_truth_binary'] is not None)
        malicious = sum(1 for p in prompts if p['ground_truth_binary'] is True)
        benign = sum(1 for p in prompts if p['ground_truth_binary'] is False)
        
        print(f"Total Prompts: {total:,}")
        print(f"With Ground Truth: {with_gt:,} ({with_gt/total*100:.1f}%)")
        if with_gt > 0:
            print(f"Malicious: {malicious:,} ({malicious/with_gt*100:.1f}%)")
            print(f"Benign: {benign:,} ({benign/with_gt*100:.1f}%)")
            print(f"Class Balance Ratio: {benign/max(malicious, 1):.2f}")
        
        # Schema information
        if hasattr(self, 'dataset_info'):
            print(f"\nDetected Schema: {self.dataset_info.get('detected_schema', 'unknown')}")
            print(f"Schema Confidence: {self.dataset_info.get('confidence', 0):.2f}")
        
        # Prompt characteristics
        if NUMPY_AVAILABLE and prompts:
            lengths = [p['prompt_length'] for p in prompts]
            complexities = [p['prompt_complexity'] for p in prompts]
            
            print(f"\nPrompt Statistics:")
            print(f"  Average Length: {np.mean(lengths):.1f} ± {np.std(lengths):.1f} characters")
            print(f"  Length Range: {np.min(lengths)} - {np.max(lengths)}")
            print(f"  Average Complexity: {np.mean(complexities):.3f}")
        
        print("="*80)

    def _print_final_summary(self, total_results: int, output_files: Dict[str, str]):
        """Print final processing summary"""
        print("\n" + "="*80)
        print("PROCESSING COMPLETE")
        print("="*80)
        print(f"Total Results: {total_results:,}")
        print(f"Success Rate: {self.successful_requests/self.total_requests*100:.1f}%")
        if self.start_time:
            total_time = time.time() - self.start_time
            print(f"Processing Time: {total_time/60:.1f} minutes")
            print(f"Processing Rate: {total_results/(total_time/60):.1f} prompts/minute")
        
        print(f"\nOutput Directory: {self.output_directory.absolute()}")
        print("Generated Files:")
        for file_type, file_path in output_files.items():
            print(f"  {file_type.title()}: {Path(file_path).name}")
        
        print("="*80)

    def _generate_performance_plots(self, output_files: Dict[str, str]):
        """Generate performance visualization plots"""
        if not PLOTTING_AVAILABLE:
            logger.warning("Matplotlib not available - skipping plot generation")
            return
        
        try:
            # Implementation would generate various performance plots
            # This is a placeholder for the plotting functionality
            logger.info("Performance plots would be generated here (feature placeholder)")
            
        except Exception as e:
            logger.error(f"Error generating plots: {str(e)}")

def parse_enhanced_arguments():
    """Parse enhanced command line arguments"""
    parser = argparse.ArgumentParser(
        description="Azure Content Safety - Enhanced Robust Multi-Dataset Processor v5.0",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
Enhanced Features:
  • Intelligent schema detection for multiple dataset types
  • Advanced ground truth mapping with confidence scores
  • Comprehensive performance metrics with statistical analysis
  • Multiple output formats (CSV, JSON, summary report)
  • Enhanced error handling and progress tracking
  • Business impact analysis and quality indicators

Supported Dataset Types:
  • JBB Benign Behaviors Dataset
  • Jailbreak Balanced Classification Dataset
  • Mini JailBreak Dataset (28K)
  • Red Team Evaluation Datasets
  • Safety Evaluation Datasets
  • Custom datasets with auto-detection

Examples:

  # Auto-detect everything (recommended)
  python enhanced_processor.py --input dataset.csv --endpoint URL --api-key KEY

  # JBB Benign Behaviors
  python enhanced_processor.py --input jbb_benignbehaviors.csv --prompt-column "Behavior"

  # Jailbreak Dataset with custom settings
  python enhanced_processor.py --input jailbreak_balanced.csv --severity-threshold 1 --concurrent 10

  # High-throughput processing
  python enhanced_processor.py --input large_dataset.csv --concurrent 15 --rate-limit 100

  # Analysis only (no API calls)
  python enhanced_processor.py --input dataset.csv --analyze-only

  # Comprehensive analysis with plots
  python enhanced_processor.py --input dataset.csv --generate-plots --verbose

Environment Variables:
  AZURE_CONTENT_SAFETY_ENDPOINT - Azure endpoint URL
  AZURE_CONTENT_SAFETY_KEY - Azure API key
        """
    )
    
    # Required arguments
    parser.add_argument('--input', '-i', type=str, required=True,
                       help='Input CSV file path')
    
    # Authentication
    parser.add_argument('--endpoint', '-e', type=str,
                       help='Azure Content Safety endpoint URL')
    parser.add_argument('--api-key', '-k', type=str,
                       help='Azure Content Safety API key')
    
    # Column specification
    parser.add_argument('--prompt-column', '-p', type=str,
                       help='Prompt column name (auto-detects if not specified)')
    parser.add_argument('--ground-truth-column', '-g', type=str,
                       help='Ground truth column name (auto-detects if not specified)')
    
    # Processing parameters
    parser.add_argument('--concurrent', '-c', type=int, default=5,
                       help='Max concurrent requests (default: 5)')
    parser.add_argument('--rate-limit', '-r', type=int, default=30,
                       help='Rate limit per minute (default: 30)')
    parser.add_argument('--severity-threshold', type=int, default=2,
                       help='Severity threshold for blocking (0-6, default: 2)')
    parser.add_argument('--timeout', type=int, default=30,
                       help='Request timeout in seconds (default: 30)')
    
    # Output options
    parser.add_argument('--output-dir', '-o', type=str, default='azure_results',
                       help='Output directory')
    parser.add_argument('--output-prefix', type=str,
                       help='Custom output file prefix')
    
    # Analysis options
    parser.add_argument('--analyze-only', action='store_true',
                       help='Analyze dataset without API calls')
    parser.add_argument('--schema-detection', action='store_true', default=True,
                       help='Enable automatic schema detection (default: True)')
    parser.add_argument('--no-schema-detection', dest='schema_detection', action='store_false',
                       help='Disable automatic schema detection')
    
    # Utility options
    parser.add_argument('--list-columns', action='store_true',
                       help='List available columns and exit')
    parser.add_argument('--dry-run', action='store_true',
                       help='Test endpoint and file loading without processing')
    parser.add_argument('--generate-plots', action='store_true',
                       help='Generate performance plots (requires matplotlib)')
    
    # Logging and debugging
    parser.add_argument('--verbose', '-v', action='store_true',
                       help='Verbose logging')
    parser.add_argument('--debug', action='store_true',
                       help='Debug level logging')
    parser.add_argument('--quiet', '-q', action='store_true',
                       help='Minimal logging')
    
    return parser.parse_args()

async def main():
    """Enhanced main function with comprehensive error handling"""
    
    try:
        args = parse_enhanced_arguments()
    except Exception as e:
        logger.error(f"Argument parsing failed: {str(e)}")
        sys.exit(1)
    
    # Configure logging level
    if args.debug:
        logging.getLogger().setLevel(logging.DEBUG)
    elif args.verbose:
        logging.getLogger().setLevel(logging.INFO)
    elif args.quiet:
        logging.getLogger().setLevel(logging.WARNING)
    
    # Get credentials
    endpoint_url = args.endpoint or os.getenv("AZURE_CONTENT_SAFETY_ENDPOINT")
    api_key = args.api_key or os.getenv("AZURE_CONTENT_SAFETY_KEY")
    
    # Validate requirements
    if not endpoint_url and not args.analyze_only and not args.list_columns:
        logger.error("Azure endpoint URL required (use --endpoint or set AZURE_CONTENT_SAFETY_ENDPOINT)")
        sys.exit(1)
    
    if not api_key and not args.analyze_only and not args.dry_run and not args.list_columns:
        logger.error("Azure API key required (use --api-key or set AZURE_CONTENT_SAFETY_KEY)")
        sys.exit(1)
    
    # Initialize processor
    try:
        processor = EnhancedAzureContentSafetyProcessor(
            endpoint_url=endpoint_url or "https://placeholder.cognitiveservices.azure.com",
            api_key=api_key or "placeholder",
            max_concurrent_requests=args.concurrent,
            rate_limit_per_minute=args.rate_limit,
            timeout_seconds=args.timeout,
            output_directory=args.output_dir,
            prompt_column=args.prompt_column,
            ground_truth_column=args.ground_truth_column,
            severity_threshold=args.severity_threshold,
            auto_detect_schema=args.schema_detection
        )
        
        logger.info("Enhanced Azure Content Safety Processor v5.0 initialized")
        
    except Exception as e:
        logger.error(f"Failed to initialize processor: {str(e)}")
        sys.exit(1)
    
    try:
        # Handle utility options
        if args.list_columns:
            processor._handle_list_columns(args.input)
            return
        
        # Load and analyze dataset
        logger.info(f"Loading and analyzing dataset: {args.input}")
        prompts = processor.load_and_analyze_dataset(args.input)
        
        if not prompts:
            logger.error("No valid prompts loaded")
            sys.exit(1)
        
        logger.info(f"Successfully loaded {len(prompts)} prompts")
        
        # Handle analysis-only mode
        if args.analyze_only:
            logger.info("Analysis-only mode - skipping API processing")
            processor._print_dataset_analysis_summary(prompts)
            return
        
        # Handle dry run
        if args.dry_run:
            logger.info("Dry run mode - testing endpoint...")
            success = await processor.discover_working_endpoint()
            if success:
                logger.info("✅ Dry run successful!")
                processor._print_dataset_analysis_summary(prompts)
            else:
                logger.error("❌ Endpoint discovery failed")
                sys.exit(1)
            return
        
        # Process batch
        logger.info("Starting batch processing...")
        results = await processor.process_batch(prompts)
        
        # Save comprehensive results
        input_filename = Path(args.input).name
        output_files = processor.save_comprehensive_results(
            output_file=args.output_prefix,
            input_filename=input_filename
        )
        
        # Generate plots if requested
        if args.generate_plots and PLOTTING_AVAILABLE:
            processor._generate_performance_plots(output_files)
        
        # Final summary
        processor._print_final_summary(len(results), output_files)
        
    except KeyboardInterrupt:
        logger.info("Processing interrupted by user")
        sys.exit(1)
    except Exception as e:
        logger.error(f"Processing error: {str(e)}")
        if args.verbose or args.debug:
            traceback.print_exc()
        sys.exit(1)

if __name__ == "__main__":
    try:
        run_async_main(main())
    except KeyboardInterrupt:
        print("\nOperation cancelled by user")
        sys.exit(1)
    except Exception as e:
        logger.error(f"Fatal error: {str(e)}")
        sys.exit(1)
